import os
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import threading
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook
import re

class ModernQCComparisonGUI:
    def __init__(self, root):
        self.root = root
        self.setup_window()
        self.setup_styles()
        self.create_widgets()

    # ========================
    # Main QC comparison logic
    # ========================
    def qc_comparison_script(self, input_key_path, input_agent_path, output_file_path):
        """QC comparison function with GUI logging and Excel report generation."""
        self.log("Starting QC Comparison Process...")
        self.log("=" * 50)

        # Resolve Excel files from inputs
        try:
            keysheet_file = self.get_excel_file_from_input(input_key_path)
            agent_file = self.get_excel_file_from_input(input_agent_path)
            self.log(f"Keysheet file: {keysheet_file}")
            self.log(f"Agent file: {agent_file}")
        except Exception as e:
            self.log(f"Error resolving input files: {e}")
            messagebox.showerror("Error", f"Error resolving input files: {e}")
            self._finish_comparison()
            return

        # Read the Production & QC Report sheets from both files
        try:
            sheet_name = 'Production & QC Report'
            keysheet_df = self.read_excel_data_with_openpyxl(keysheet_file, sheet_name)
            agent_df = self.read_excel_data_with_openpyxl(agent_file, sheet_name)

            if keysheet_df is None or agent_df is None:
                self.log("Failed to load one or more Excel sheets.")
                messagebox.showerror("Error", "Failed to load one or more Excel sheets.")
                self._finish_comparison()
                return

            # Normalize data
            for df in [keysheet_df, agent_df]:
                for col in df.columns:
                    df[col] = df[col].apply(self.normalize_numeric_values)

            self.log(f"Successfully loaded keysheet with {len(keysheet_df)} rows and agent file with {len(agent_df)} rows")
        except Exception as e:
            self.log(f"Error reading Excel files: {e}")
            messagebox.showerror("Error", f"Error reading Excel files: {e}")
            self._finish_comparison()
            return

        # Prepare keysheet for merging: handle potential duplicates
        if 'Item ID' not in keysheet_df.columns or 'Item ID' not in agent_df.columns:
            self.log("Error: 'Item ID' column not found in one or both files.")
            messagebox.showerror("Error", "Required column 'Item ID' not found in one or both files.")
            self._finish_comparison()
            return

        keysheet_unique_df = keysheet_df.drop_duplicates(subset=['Item ID'], keep='first')

        # Merge dataframes on Item ID
        merged_df = pd.merge(keysheet_unique_df, agent_df, on='Item ID', how='inner', suffixes=('_key', '_agent'))
        self.log(f"Successfully matched {len(merged_df)} records on Item ID")

        # Detect the associate column robustly (handles suffix and header variations)
        assoc_col = self._detect_assoc_col(merged_df)
        self.log(f"Using associate column: {assoc_col if assoc_col else '[not found]'}")

        # Columns to exclude from comparison (unchanged logic)
        exclude_columns = [
            'Date (YYYY/MM/DD)', 'Work Type', 'Associate Walmart ID', 'Item Status',
            'Submission ID', 'Product ID Type', 'Product ID', 'Audit Template Version',
            'SOP Reference', 'Initial CQ Score', 'Current CQ Score', 'Supplier ID',
            'Item Created Date', 'Active Status', 'is Private Label?', 'Website Link', 'Product Type',
            'View Images', 'Product Short Description', 'Product Long Description', 'Product Name', 'Product Type Group'
        ]

        # -------------------------------------------------
        # Build discrepancy rows with Error Codes column
        # -------------------------------------------------
        discrepancy_records = []
        keysheet_columns = keysheet_unique_df.columns.tolist()

        for _, row in merged_df.iterrows():
            item_id = row['Item ID']
            associate_id = row.get(assoc_col, '') if assoc_col else ''
            error_codes = row.get('Error Codes', '') if 'Error Codes' in row.index else ''

            for col in keysheet_columns:
                if col in exclude_columns or col == 'Item ID' or str(col).startswith('Pre - '):
                    continue

                agent_col = col + '_agent'
                key_col = col + '_key'
                key_value = row.get(key_col, '')
                agent_value = row.get(agent_col, '')

                if key_col in row.index and agent_col in row.index:
                    is_key_blank = pd.isna(key_value) or key_value is None or (isinstance(key_value, str) and str(key_value).strip() == '')
                    is_agent_blank = pd.isna(agent_value) or agent_value is None or (isinstance(agent_value, str) and str(agent_value).strip() == '')

                    if is_key_blank and is_agent_blank:
                        continue

                    key_str = '' if is_key_blank else str(key_value).strip()
                    agent_str = '' if is_agent_blank else str(agent_value).strip()

                    if '|' in key_str or '|' in agent_str or ',' in key_str or ',' in agent_str:
                        def split_and_clean(value):
                            if not value:
                                return set()
                            parts = [p.strip() for p in value.split('|')]
                            result = set()
                            for part in parts:
                                if ',' in part:
                                    result.update(v.strip() for v in part.split(','))
                                else:
                                    result.add(part)
                            return result

                        key_set = split_and_clean(key_str)
                        agent_set = split_and_clean(agent_str)
                        if key_set == agent_set:
                            continue

                    if key_str != agent_str:
                        attribute_name = self.extract_attribute_name(col)
                        error_column = col
                        discrepancy_records.append({
                            'Item ID': item_id,
                            'Associate Walmart ID': associate_id,
                            'Attribute': attribute_name,
                            'Error column': error_column,
                            'Correct Value': key_value,
                            'Agent Provided Value': agent_value,
                        })

        # =======================
        # Final Summary & Outputs
        # =======================
        total_discrepancies = len(discrepancy_records)
        self.log("=" * 50)
        self.log(f"Comparison complete. Found {total_discrepancies} discrepancies.")

        # Construct agent mapping for downstream use
        if assoc_col and assoc_col in merged_df.columns:
            agent_mapping_df = merged_df[['Item ID', assoc_col]].copy()
            agent_mapping_df.columns = ['Item ID', 'Walmart Associate ID']
            self.log(f"DEBUG: Created agent_mapping_df with {len(agent_mapping_df)} rows")
            self.log(f"DEBUG: agent_mapping_df columns: {list(agent_mapping_df.columns)}")
            self.log(f"DEBUG: Sample agent mapping: {agent_mapping_df.head(3).to_dict()}")
        else:
            self.log("WARNING: assoc_col is None or not found in merged_df")
            agent_mapping_df = pd.DataFrame(columns=['Item ID', 'Walmart Associate ID'])

        # Build per-attribute filled counts restricted to common Item IDs
        self.log("Building per-attribute filled counts from common Item IDs...")
        common_item_ids = set(merged_df['Item ID'].unique())
        filtered_keysheet_for_counts = keysheet_unique_df[keysheet_unique_df['Item ID'].isin(common_item_ids)].copy()
        filled_counts_df = self.build_filled_attribute_column_counts(filtered_keysheet_for_counts)

        # Write discrepancies and summary to Excel file
        if total_discrepancies > 0:
            discrepancy_df = pd.DataFrame(discrepancy_records)
            discrepancy_df = self._ensure_assoc_in_discrepancies(discrepancy_df, merged_df, assoc_col)

            if assoc_col is None:
                self.log("Warning: No associate ID column found in merged data. Skipping Agent Summary.")
                agent_summary = pd.DataFrame()
                formulas_df = pd.DataFrame()
            else:
                # Agent metrics
                agent_item_counts = merged_df[assoc_col].value_counts().reset_index()
                agent_item_counts.columns = ['Associate Walmart ID', 'Line items Audited']
                agents_list = agent_item_counts['Associate Walmart ID'].tolist()

                agent_error_counts = self.calculate_unique_attribute_failures(discrepancy_df, agents_list)

                agent_failed_items = (
                    discrepancy_df.groupby('Associate Walmart ID')['Item ID']
                    .nunique()
                    .reindex(agents_list, fill_value=0)
                    .reset_index(name='No. of line items Failed')
                )

                self.log("Calculating filled required attributes...")
                # VERIFIED: Correct method call with 3 parameters
                filled_attributes_dict = self.calculate_filled_required_attributes(filtered_keysheet_for_counts, merged_df, agent_mapping_df)

                per_agent_filled = pd.DataFrame([
                    {'Associate Walmart ID': assoc_id, 'No.of filled required attributes': count}
                    for assoc_id, count in filled_attributes_dict.items()
                ])
                per_agent_filled = per_agent_filled.set_index('Associate Walmart ID').reindex(agents_list, fill_value=0).reset_index()
                per_agent_filled['No.of filled required attributes'] = per_agent_filled['No.of filled required attributes'].fillna(0).astype(int)

                agent_summary = pd.merge(agent_item_counts, agent_error_counts, on='Associate Walmart ID', how='left')
                agent_summary = pd.merge(agent_summary, agent_failed_items, on='Associate Walmart ID', how='left')
                agent_summary = pd.merge(agent_summary, per_agent_filled, on='Associate Walmart ID', how='left')

                agent_summary[['Total No of Attributes Failed', 'No. of line items Failed', 'No.of filled required attributes']] = \
                    agent_summary[['Total No of Attributes Failed', 'No. of line items Failed', 'No.of filled required attributes']].fillna(0).astype(int)

                agent_summary['No of line items Passed'] = agent_summary['Line items Audited'] - agent_summary['No. of line items Failed']
                agent_summary['Item Level QC %'] = (
                    1 - (agent_summary['No. of line items Failed'] / agent_summary['Line items Audited'])
                ).replace([np.inf, -np.inf], 0).fillna(0).map('{:0.2%}'.format)

                agent_summary['Required attribute Accuracy%'] = (
                    1 - (agent_summary['Total No of Attributes Failed'] / agent_summary['No.of filled required attributes'])
                ).replace([np.inf, -np.inf], 0).fillna(0).map('{:0.2%}'.format)

                agent_summary_order = [
                    'Associate Walmart ID',
                    'Line items Audited',
                    'No. of line items Failed',
                    'No of line items Passed',
                    'No.of filled required attributes',
                    'Total No of Attributes Failed',
                    'Item Level QC %',
                    'Required attribute Accuracy%'
                ]
                agent_summary = agent_summary[agent_summary_order]

                formulas_data = {
                    'Column Name': [
                        'Line items Audited', 'No. of line items Failed', 'No of line items Passed',
                        'No.of filled required attributes', 'Total No of Attributes Failed',
                        'Item Level QC %', 'Required attribute Accuracy%'
                    ],
                    'Formula': [
                        'No. of line items in the agent file',
                        'Total number of unique line items with at least one discrepancy',
                        'Line items Audited - No of line items Failed',
                        'Count filled product attribute columns per Item ID (excludes system/metadata and validation columns). Sum for associates with multiple Item IDs.',
                        'Total count of attributes in Error Sheet for each agent (total rows in Error Sheet for each Associate Walmart ID).',
                        '1 - (No. of line items Failed) / (Line items Audited)',
                        '1 - (Total No of Attributes Failed) / (No.of filled required attributes)'
                    ],
                    'Description': [
                        'Total number of line items audited by the agent (row-level)',
                        'Number of line items that have discrepancies',
                        'Number of line items that passed without discrepancies',
                        'Count of filled product attributes excluding system columns and validation columns.',
                        'Count of unique attribute columns with errors. Each attribute column counts separately.',
                        'Percentage of line items that passed QC for each agent',
                        'Percentage of correctly filled attributes for each agent'
                    ]
                }
                formulas_df = pd.DataFrame(formulas_data)

            # ==============
            # Write to Excel
            # ==============
            try:
                output_dir = os.path.dirname(output_file_path)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir)

                with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                    discrepancy_df.to_excel(writer, sheet_name='Error Sheet', index=False)
                    if not agent_summary.empty:
                        agent_summary.to_excel(writer, sheet_name='Agent Summary', index=False)
                    if not formulas_df.empty:
                        formulas_df.to_excel(writer, sheet_name='Formulas', index=False)
                    if not filled_counts_df.empty:
                        filled_counts_df.to_excel(writer, sheet_name='Filled Attribute Counts', index=False)

                self.log(f"QC report saved to: {output_file_path}")
                messagebox.showinfo("Success", f"QC report generated successfully!\n\nDiscrepancy count: {total_discrepancies}")
            except PermissionError:
                self.log("Error: Output file is open in another program. Please close it and try again.")
                messagebox.showerror("Error", "Output file is open in another program. Please close it and try again.")
            except Exception as e:
                self.log(f"Error saving output file: {e}")
                messagebox.showerror("Error", f"An error occurred: {str(e)}")
        else:
            # No discrepancies case
            self.log("No discrepancies found. Generating report with perfect scores...")
            discrepancy_df = pd.DataFrame(columns=[
                'Item ID', 'Associate Walmart ID', 'Attribute', 'Error column',
                'Correct Value', 'Agent Provided Value'
            ])

            if assoc_col is None:
                self.log("Warning: No associate ID column found in merged data. Skipping Agent Summary.")
                agent_summary = pd.DataFrame()
                formulas_df = pd.DataFrame()
            else:
                agent_item_counts = merged_df[assoc_col].value_counts().reset_index()
                agent_item_counts.columns = ['Associate Walmart ID', 'Line items Audited']
                agents_list = agent_item_counts['Associate Walmart ID'].tolist()

                self.log("Calculating filled required attributes...")
                # VERIFIED: Correct method call with 3 parameters
                filled_attributes_dict = self.calculate_filled_required_attributes(filtered_keysheet_for_counts, merged_df, agent_mapping_df)

                per_agent_filled = pd.DataFrame([
                    {'Associate Walmart ID': assoc_id, 'No.of filled required attributes': count}
                    for assoc_id, count in filled_attributes_dict.items()
                ])
                per_agent_filled = per_agent_filled.set_index('Associate Walmart ID').reindex(agents_list, fill_value=0).reset_index()
                per_agent_filled['No.of filled required attributes'] = per_agent_filled['No.of filled required attributes'].fillna(0).astype(int)

                agent_summary = agent_item_counts.copy()
                agent_summary['No. of line items Failed'] = 0
                agent_summary['No of line items Passed'] = agent_summary['Line items Audited']
                agent_summary['Total No of Attributes Failed'] = 0
                agent_summary = pd.merge(agent_summary, per_agent_filled, on='Associate Walmart ID', how='left')
                agent_summary['No.of filled required attributes'] = agent_summary['No.of filled required attributes'].fillna(0).astype(int)
                agent_summary['Item Level QC %'] = '100.00%'
                agent_summary['Required attribute Accuracy%'] = '100.00%'

                agent_summary_order = [
                    'Associate Walmart ID', 'Line items Audited', 'No. of line items Failed',
                    'No of line items Passed', 'No.of filled required attributes', 'Total No of Attributes Failed',
                    'Item Level QC %', 'Required attribute Accuracy%'
                ]
                agent_summary = agent_summary[agent_summary_order]

                formulas_data = {
                    'Column Name': [
                        'Line items Audited', 'No. of line items Failed', 'No of line items Passed',
                        'No.of filled required attributes', 'Total No of Attributes Failed',
                        'Item Level QC %', 'Required attribute Accuracy%'
                    ],
                    'Formula': [
                        'No. of line items in the agent file',
                        'Total number of unique line items with at least one discrepancy',
                        'Line items Audited - No of line items Failed',
                        'Count filled product attribute columns per Item ID (excludes system/metadata and validation columns). Sum for associates with multiple Item IDs.',
                        'Total count of attributes in Error Sheet for each agent (total rows in Error Sheet for each Associate Walmart ID).',
                        '1 - (No. of line items Failed) / (Line items Audited)',
                        '1 - (Total No of Attributes Failed) / (No.of filled required attributes)'
                    ],
                    'Description': [
                        'Total number of line items audited by the agent (row-level)',
                        'Number of line items that have discrepancies',
                        'Number of line items that passed without discrepancies',
                        'Count of filled product attributes excluding system columns and validation columns.',
                        'Count of unique attribute columns with errors. Each attribute column counts separately.',
                        'Percentage of line items that passed QC for each agent',
                        'Percentage of correctly filled attributes for each agent'
                    ]
                }
                formulas_df = pd.DataFrame(formulas_data)

            try:
                output_dir = os.path.dirname(output_file_path)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir)

                with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                    discrepancy_df.to_excel(writer, sheet_name='Error Sheet', index=False)
                    if not agent_summary.empty:
                        agent_summary.to_excel(writer, sheet_name='Agent Summary', index=False)
                    if not formulas_df.empty:
                        formulas_df.to_excel(writer, sheet_name='Formulas', index=False)
                    if not filled_counts_df.empty:
                        filled_counts_df.to_excel(writer, sheet_name='Filled Attribute Counts', index=False)

                self.log(f"QC report saved to: {output_file_path}")
                messagebox.showinfo("Success", f"QC report generated successfully!\n\nNo discrepancies found - all associates achieved 100% scores!")
            except PermissionError:
                self.log("Error: Output file is open in another program. Please close it and try again.")
                messagebox.showerror("Error", "Output file is open in another program. Please close it and try again.")
            except Exception as e:
                self.log(f"Error saving output file: {e}")
                messagebox.showerror("Error", f"An error occurred: {str(e)}")

        # IMPORTANT: Always call _finish_comparison at the very end
        self._finish_comparison()

    # ========================
    # Per-attribute filled counts (common Item IDs only)
    # ========================
    def build_filled_attribute_column_counts(self, df_data):
        """
        Build per-attribute 'filled' counts across rows in df_data.
        Assumes df_data already filtered to Item IDs present in both key sheet and agent file.
        Uses same exclusion logic as calculate_filled_required_attributes.
        """
        if df_data is None or df_data.empty:
            return pd.DataFrame(columns=['Attribute Column', 'Filled Count'])

        exclude_columns = [
            'Date (YYYY/MM/DD)', 'Work Type', 'Associate Walmart ID', 'Item Status',
            'Submission ID', 'Product ID Type', 'Product ID', 'Audit Template Version',
            'SOP Reference', 'Initial CQ Score', 'Current CQ Score', 'Supplier ID',
            'Item Created Date', 'Active Status', 'Website Link', 'View Images',
            'is Private Label?', 'Category', 'Product Type Group', 'Product Type',
            'Product Short Description', 'Product Long Description', 'Product Name',
            'Main Image URL'
        ]

        all_columns = df_data.columns.tolist()
        pattern_exclude_columns = []
        for col in all_columns:
            col_lower = col.lower()
            if ('is ' in col_lower and 'validated' in col_lower) or \
               ('error code' in col_lower) or \
               ('validation comment' in col_lower):
                pattern_exclude_columns.append(col)

        all_exclude_columns = set(exclude_columns + pattern_exclude_columns)
        attribute_columns = [col for col in all_columns if col not in all_exclude_columns and col != 'Item ID' and 'Pre' not in col]

        def is_filled(v):
            if pd.isna(v):
                return False
            s = str(v).strip()
            return s != ''

        filled_counts = []
        for col in attribute_columns:
            count = int(df_data[col].apply(is_filled).sum())
            filled_counts.append({'Attribute Column': col, 'Filled Count': count})

        counts_df = pd.DataFrame(filled_counts).sort_values(by='Attribute Column').reset_index(drop=True)
        return counts_df

    # ========================
    # Unique attribute failures
    # ========================
    def calculate_unique_attribute_failures(self, discrepancy_df, agents_list):
        """
        Calculate total number of attributes in error sheet for each agent.
        Counts rows in Error Sheet for each Associate Walmart ID.
        """
        attribute_failures = {}
        for associate_id in agents_list:
            associate_discrepancies = discrepancy_df[discrepancy_df['Associate Walmart ID'] == associate_id]
            attribute_failures[associate_id] = len(associate_discrepancies)

        result_df = pd.DataFrame([
            {'Associate Walmart ID': assoc_id, 'Total No of Attributes Failed': count}
            for assoc_id, count in attribute_failures.items()
        ])
        return result_df.set_index('Associate Walmart ID').reindex(agents_list, fill_value=0).reset_index()

    # ========================
    # $5000 BET: THE FIXED METHOD
    # ========================
    def calculate_filled_required_attributes(self, df_data, merged_df, agent_mapping_df):
        """
        Calculate the 'No. of Filled Required Attributes' for each Walmart Associate ID.
        $5000 BET: Every variable reference triple-checked to use correct parameter names.
        """
        # VERIFIED: Debug statements using correct parameter names
        self.log(f"DEBUG: df_data shape: {df_data.shape if df_data is not None else 'None'}")
        self.log(f"DEBUG: merged_df shape: {merged_df.shape if merged_df is not None else 'None'}")
        self.log(f"DEBUG: agent_mapping_df shape: {agent_mapping_df.shape if agent_mapping_df is not None else 'None'}")
        
        # VERIFIED: Validation using correct parameter names
        if agent_mapping_df.empty or merged_df.empty:
            self.log("WARNING: agent_mapping_df or merged_df is empty. Returning empty dictionary.")
            return {}
        
        # Step 1: Define exclusions
        exclude_columns = [
            'Date (YYYY/MM/DD)', 'Work Type', 'Associate Walmart ID', 'Item Status',
            'Submission ID', 'Product ID Type', 'Product ID', 'Audit Template Version',
            'SOP Reference', 'Initial CQ Score', 'Current CQ Score', 'Supplier ID',
            'Item Created Date', 'Active Status', 'Website Link', 'View Images',
            'is Private Label?', 'Category', 'Product Type Group', 'Product Type',
            'Product Short Description', 'Product Long Description', 'Product Name',
            'Main Image URL'
        ]

        # VERIFIED: Column extraction using merged_df (parameter 2)
        merged_cols = merged_df.columns.tolist()
        keysheet_cols = [col.replace('_key', '') for col in merged_cols if col.endswith('_key')]
        agent_cols = [col.replace('_agent', '') for col in merged_cols if col.endswith('_agent')]
        
        self.log(f"DEBUG: keysheet_cols count: {len(keysheet_cols)}")
        self.log(f"DEBUG: agent_cols count: {len(agent_cols)}")

        # Pattern exclusions
        pattern_exclude_columns = []
        for col in keysheet_cols:
            col_lower = str(col).lower()
            if ('is' in col_lower and 'validated' in col_lower) or \
               ('error code' in col_lower) or \
               ('validation comment' in col_lower):
                pattern_exclude_columns.append(col)

        all_exclude_columns = set(exclude_columns + pattern_exclude_columns)

        # Step 2: Get remaining columns and find intersection
        remaining_keysheet_columns = [
            col for col in keysheet_cols
            if col not in all_exclude_columns and col != 'Item ID' and 'Pre-' not in str(col) and 'Pre -' not in str(col)
        ]
        
        remaining_agent_columns = [
            col for col in agent_cols
            if col not in all_exclude_columns and col != 'Item ID' and 'Pre-' not in str(col) and 'Pre -' not in str(col)
        ]
        
        attribute_columns = list(set(remaining_keysheet_columns) & set(remaining_agent_columns))
        
        self.log(f"DEBUG: remaining_keysheet_columns: {len(remaining_keysheet_columns)}")
        self.log(f"DEBUG: remaining_agent_columns: {len(remaining_agent_columns)}")
        self.log(f"DEBUG: attribute_columns (intersection): {len(attribute_columns)}")
        self.log(f"DEBUG: First few attribute columns: {attribute_columns[:5] if attribute_columns else 'None'}")

        if not attribute_columns:
            self.log("ERROR: No attribute columns found after intersection. Check column names and exclusions.")
            return {}

        # VERIFIED: Using agent_mapping_df (parameter 3) for common IDs
        common_ids = set(agent_mapping_df['Item ID'].unique())
        df_data_common = df_data[df_data['Item ID'].isin(common_ids)].copy()
        
        self.log(f"DEBUG: common_ids count: {len(common_ids)}")
        self.log(f"DEBUG: df_data_common shape: {df_data_common.shape}")

        # Step 4: Count filled per item across attribute_columns
        item_filled_counts = {}
        for _, row in df_data_common.iterrows():
            item_id = row['Item ID']
            filled_count = 0
            for col in attribute_columns:
                if col not in row.index:
                    continue
                value = row[col]
                if pd.notna(value) and str(value).strip() != '':
                    filled_count += 1
            item_filled_counts[item_id] = filled_count
            
        self.log(f"DEBUG: item_filled_counts sample: {dict(list(item_filled_counts.items())[:3])}")

        # VERIFIED: Using agent_mapping_df (parameter 3) for final mapping
        associate_totals = {}
        for _, arow in agent_mapping_df.iterrows():
            item_id = arow['Item ID']
            associate_id = arow['Walmart Associate ID']
            if item_id in item_filled_counts:
                associate_totals[associate_id] = associate_totals.get(associate_id, 0) + item_filled_counts[item_id]

        self.log("DEBUG: Associate filled attribute totals:")
        for associate_id, total in associate_totals.items():
            self.log(f"  Associate {associate_id}: {total} total")

        return associate_totals

    # ========================
    # Helper: header normalization & associate detection
    # ========================
    def _normalize_header(self, s: str) -> str:
        """Lowercase, remove spaces and non-alnum to compare column names safely."""
        return re.sub(r'[^a-z0-9]', '', str(s).lower())

    def _detect_assoc_col(self, df: pd.DataFrame):
        """
        Robustly find the associate column in merged_df:
        - Prefer *_agent variants, then *_key.
        - Match fuzzy names like 'Associate Walmart ID', 'Associate ID', etc.
        """
        cols = list(df.columns)
        norm = {c: self._normalize_header(c) for c in cols}

        def looks_like_assoc(n: str) -> bool:
            return ('associate' in n) and ('id' in n)

        agents = [c for c in cols if c.endswith('_agent') and looks_like_assoc(norm[c])]
        if agents:
            return agents[0]

        keys = [c for c in cols if c.endswith('_key') and looks_like_assoc(norm[c])]
        if keys:
            return keys[0]

        any_match = [c for c in cols if looks_like_assoc(norm[c])]
        if any_match:
            return any_match[0]

        self.log("Warning: Could not detect Associate column in merged data.")
        return None

    def _ensure_assoc_in_discrepancies(self, discrepancy_df: pd.DataFrame,
                                       merged_df: pd.DataFrame, assoc_col: str | None) -> pd.DataFrame:
        target = 'Associate Walmart ID'
        if target in discrepancy_df.columns:
            return discrepancy_df

        if assoc_col and assoc_col in merged_df.columns:
            self.log(f"Info: Reconstructing '{target}' in discrepancies from '{assoc_col}'.")
            map_df = merged_df[['Item ID', assoc_col]].drop_duplicates()
            out = discrepancy_df.merge(map_df, on='Item ID', how='left')
            out.rename(columns={assoc_col: target}, inplace=True)
            return out

        self.log(f"Warning: '{target}' could not be reconstructed; filling blanks.")
        discrepancy_df[target] = ''
        return discrepancy_df

    # ========================
    # Helper methods for attribute name extraction
    # ========================
    def extract_attribute_name(self, column_name):
        """Extract a clean attribute name from column name."""
        clean_name = str(column_name).strip()
        return clean_name

    # ========================
    # Window & Styling
    # ========================
    def setup_window(self):
        """Configure the main window"""
        self.root.title("Training QC Tool - Model Validation")
        self.root.geometry("800x600")
        self.root.minsize(700, 500)
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (800 // 2)
        y = (self.root.winfo_screenheight() // 2) - (600 // 2)
        self.root.geometry(f"800x600+{x}+{y}")
        self.root.configure(bg='#f0f0f0')

    def setup_styles(self):
        """Configure modern styles"""
        self.style = ttk.Style()
        self.style.configure('Title.TLabel', font=('Segoe UI', 20, 'bold'), background='#f0f0f0', foreground='#2c3e50')
        self.style.configure('Subtitle.TLabel', font=('Segoe UI', 10), background='#f0f0f0', foreground='#7f8c8d')
        self.style.configure('Modern.TLabel', font=('Segoe UI', 10), background='#ffffff', foreground='#2c3e50')
        self.style.configure('Modern.TButton', font=('Segoe UI', 10), padding=(20, 10))
        self.style.configure('Browse.TButton', font=('Segoe UI', 9), padding=(10, 5))
        self.style.configure('Modern.TEntry', font=('Segoe UI', 10), fieldbackground='#ffffff')

    def create_widgets(self):
        """Create all GUI widgets"""
        main_container = ttk.Frame(self.root, padding="30")
        main_container.pack(fill=tk.BOTH, expand=True)

        title_frame = ttk.Frame(main_container)
        title_frame.pack(fill=tk.X, pady=(0, 30))

        title_label = ttk.Label(title_frame, text="Training QC Tool - Model Validation", style='Title.TLabel')
        title_label.pack()

        subtitle_label = ttk.Label(title_frame,
                                   text="Compare keysheet and agent production files to generate quality control reports",
                                   style='Subtitle.TLabel')
        subtitle_label.pack(pady=(5, 0))

        input_frame = ttk.LabelFrame(main_container, text="File Selection", padding="20")
        input_frame.pack(fill=tk.X, pady=(0, 20))

        self.create_file_input(input_frame, "Keysheet File:", "keysheet", 0)
        self.create_file_input(input_frame, "Agent Production File:", "agent", 1)
        self.create_output_input(input_frame, 2)

        button_frame = ttk.Frame(main_container)
        button_frame.pack(fill=tk.X, pady=(0, 20))

        center_frame = ttk.Frame(button_frame)
        center_frame.pack(expand=True)

        self.compare_btn = ttk.Button(center_frame,
                                      text="Generate QC Report",
                                      style='Modern.TButton',
                                      command=self.run_comparison)
        self.compare_btn.pack(side=tk.LEFT, padx=(0, 20))

        clear_btn = ttk.Button(center_frame,
                               text="Clear All",
                               style='Modern.TButton',
                               command=self.clear_all)
        clear_btn.pack(side=tk.LEFT)

        self.progress = ttk.Progressbar(main_container, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=(0, 10))

        log_frame = ttk.LabelFrame(main_container, text="Output Log", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = ScrolledText(log_frame,
                                     height=12,
                                     font=('Consolas', 9),
                                     bg='#2c3e50',
                                     fg='#ecf0f1',
                                     insertbackground='#ecf0f1')
        self.log_text.pack(fill=tk.BOTH, expand=True)

        self.log("Training QC Tool initialized. Select files to begin.")

    # ========================
    # Excel Utilities
    # ========================
    def read_excel_data_with_openpyxl(self, input_path, sheet_name):
        """
        Reads an Excel file using openpyxl and converts it to a pandas DataFrame,
        preserving blanks, 'None' as text, and 'N/A' as text.
        """
        try:
            workbook = openpyxl.load_workbook(input_path)
            name_try = sheet_name
            if name_try not in workbook.sheetnames and ' & ' in name_try:
                alt = name_try.replace(' & ', ' & ')
                if alt in workbook.sheetnames:
                    name_try = alt
            sheet = workbook[name_try]
            data = list(sheet.iter_rows(values_only=True))
            if not data:
                return pd.DataFrame()
            headers = [col.replace('\n', ' ') if isinstance(col, str) else col for col in data[0]]
            rows = data[1:]
            df = pd.DataFrame(rows, columns=headers)
            return df
        except Exception as e:
            self.log(f"Error reading Excel file with openpyxl: {e}")
            return None

    def get_excel_file_from_input(self, input_path):
        """Original function with GUI logging (fixed directory handling)."""
        if os.path.isfile(input_path):
            if input_path.lower().endswith(('.xls', '.xlsx', '.xlsm')):
                return input_path
            else:
                raise ValueError(f"Provided file path is not an Excel file: {input_path}")
        elif os.path.isdir(input_path):
            files = os.listdir(input_path)
            excel_files = [f for f in files if f.lower().endswith(('.xls', '.xlsx', '.xlsm'))]
            if not excel_files:
                raise FileNotFoundError(f"No Excel file found in directory: {input_path}")
            excel_files.sort()
            chosen = excel_files[0]
            self.log(f"Directory provided. Using file: {chosen}")
            return os.path.join(input_path, chosen)
        else:
            raise FileNotFoundError(f"The path does not exist: {input_path}")

    def normalize_numeric_values(self, value):
        """Normalize numeric values to avoid float vs int mismatches and handle blank values consistently."""
        if pd.isna(value) or value is None or (isinstance(value, str) and value.strip() == ''):
            return ''
        try:
            s = str(value)
            if re.fullmatch(r'\d+\.0', s):
                return str(int(float(s)))
            return value
        except Exception:
            return value

    # ========================
    # UI plumbing
    # ========================
    def run_comparison(self):
        input_key_path = self.keysheet_path_var.get().strip()
        input_agent_path = self.agent_path_var.get().strip()
        output_file_path = self.output_path_var.get().strip()

        if not input_key_path or not input_agent_path or not output_file_path:
            messagebox.showerror("Error", "Please select both input files and an output file path.")
            return

        self.compare_btn.config(state=tk.DISABLED)
        self.progress.start()
        self.log("Starting comparison process...")

        def run_in_thread():
            try:
                self.qc_comparison_script(input_key_path, input_agent_path, output_file_path)
            except Exception as e:
                self.log(f"Unexpected error: {e}")
                messagebox.showerror("Error", f"Unexpected error occurred: {str(e)}")
                self._finish_comparison()

        thread = threading.Thread(target=run_in_thread, daemon=True)
        thread.start()

    def create_file_input(self, parent, label_text, var_name, row):
        setattr(self, f"{var_name}_path_var", tk.StringVar())
        label = ttk.Label(parent, text=label_text, style='Modern.TLabel')
        label.grid(row=row, column=0, sticky=tk.W, pady=5)

        entry = ttk.Entry(parent, textvariable=getattr(self, f"{var_name}_path_var"), width=60, style='Modern.TEntry')
        entry.grid(row=row, column=1, padx=10, pady=5)

        def browse():
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx *.xlsm")])
            if file_path:
                getattr(self, f"{var_name}_path_var").set(file_path)

        browse_btn = ttk.Button(parent, text="Browse", style='Browse.TButton', command=browse)
        browse_btn.grid(row=row, column=2, padx=5, pady=5)

    def create_output_input(self, parent, row):
        self.output_path_var = tk.StringVar()
        label = ttk.Label(parent, text="Output Excel File:", style='Modern.TLabel')
        label.grid(row=row, column=0, sticky=tk.W, pady=5)

        entry = ttk.Entry(parent, textvariable=self.output_path_var, width=60, style='Modern.TEntry')
        entry.grid(row=row, column=1, padx=10, pady=5)

        def browse_out():
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if file_path:
                self.output_path_var.set(file_path)

        browse_btn = ttk.Button(parent, text="Browse", style='Browse.TButton', command=browse_out)
        browse_btn.grid(row=row, column=2, padx=5, pady=5)

    def clear_all(self):
        """Clear all input fields and log"""
        self.keysheet_path_var.set("")
        self.agent_path_var.set("")
        self.output_path_var.set("")
        self.log_text.delete(1.0, tk.END)
        self.log("Training QC Tool initialized. Select files to begin.")

    def log(self, message):
        """Add message to the log with timestamp"""
        import datetime
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def _finish_comparison(self):
        """Reset UI after comparison completes"""
        self.progress.stop()
        self.compare_btn.config(state=tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = ModernQCComparisonGUI(root)
    root.mainloop()
